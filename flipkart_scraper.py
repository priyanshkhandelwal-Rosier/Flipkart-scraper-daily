import pandas as pd
from bs4 import BeautifulSoup
import re
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
from openpyxl import Workbook
from openpyxl.styles import Font

# --------------------- CONFIGURATION ---------------------
# GitHub Secrets se ID/Pass lenge
YOUR_EMAIL = os.environ.get('EMAIL_USER')
APP_PASSWORD = os.environ.get('EMAIL_PASS')

# Jisko email bhejna hai uska address yahan likhein
TO_EMAIL = "receiver@gmail.com"  

EXCEL_FILE = "flipkart_rosier_products.xlsx"
HTML_FILE = "flipkart.html"
BASE_URL = "https://www.flipkart.com"

# Security Check
if not YOUR_EMAIL or not APP_PASSWORD:
    print("Error: GitHub Secrets (EMAIL_USER / EMAIL_PASS) set nahi hain!")
    exit()

# ------------------------------------------------------

# 1. HTML Read karna
print("Reading HTML file...")
try:
    with open(HTML_FILE, 'r', encoding='utf-8') as file:
        html_content = file.read()
except FileNotFoundError:
    print(f"Error: '{HTML_FILE}' file Repository me nahi mili! Please upload karein.")
    exit()

soup = BeautifulSoup(html_content, 'html.parser')
product_cards = soup.find_all('div', attrs={"data-id": True})

print(f"Total cards found: {len(product_cards)}")

data_rows = []

for card in product_cards:
    # Title Tag dhoondna (Aapki di gayi class 'pIpigb' use ho rahi hai)
    title_tag = card.find('a', class_="pIpigb")
    # Agar specific class na mile, to generic 'a' tag try karein (Backup logic)
    if not title_tag:
        title_tag = card.find('a', title=True)

    if not title_tag:
        continue
   
    title = title_tag.get_text(strip=True)
   
    # Filter: Sirf Rosier products
    if 'rosier' not in title.lower():
        continue
   
    # Product URL Extraction
    product_url = ""
    href = title_tag.get('href')
    if href:
        product_url = href if href.startswith('http') else BASE_URL + href
   
    # Variant Extraction
    variant = "-"
    quantity_tag = card.find('div', class_="U_GKRr") # Class check kr lena HTML me same hai na
    if quantity_tag:
        variant = quantity_tag.get_text(strip=True)
    else:
        # Fallback: Title me se quantity nikalne ki koshish
        match = re.search(r'(\d+(?:\.\d+)?\s*(?:kg|g|ml|l|piece|pack|pcs))', title, re.IGNORECASE)
        if match:
            variant = match.group(1)
   
    # Price Extraction
    price = "-"
    price_tag = card.find('div', class_="hZ3P6w") # Class check kr lena
    if price_tag:
        price_text = price_tag.get_text(strip=True)
        # Currency symbol saaf karna optional hai, par rehne dete hain
        if any(word in price_text.lower() for word in ['out of stock', 'unavailable', 'sold out']):
            price = "-"
        else:
            price = price_text 
   
    # Stock Status
    stock_status = "In Stock"
    card_text = card.get_text().lower()
    if "out of stock" in card_text or "unavailable" in card_text or "sold out" in card_text:
        stock_status = "Out of Stock"
   
    data_rows.append({
        "Title": title,
        "Variant": variant,
        "Price": price,
        "Stock": stock_status,
        "URL": product_url
    })

if not data_rows:
    print("HTML mein koi Rosier product nahi mila.")
    exit()

# DataFrame sirf debugging ke liye, Excel hum openpyxl se banayenge
df = pd.DataFrame(data_rows)
print(f"Total Rosier Products: {len(df)}")

# --------------------- Excel Generation (Direct OpenPyXL) ---------------------
wb = Workbook()
ws = wb.active
ws.title = "Rosier Products"

# Header
headers = ["Title", "Variant", "Price", "Stock"]
for col_num, header in enumerate(headers, 1): 
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)

# Data Rows Filling
for row_idx, row in enumerate(data_rows, start=2):
    # Column 1: Title (With Hyperlink)
    title_cell = ws.cell(row=row_idx, column=1)
    
    if row['URL']:
        title_cell.value = row['Title']
        title_cell.hyperlink = row['URL']
        title_cell.font = Font(color="0000FF", underline="single") # Blue & Underline
    else:
        title_cell.value = row['Title']

    # Other Columns
    ws.cell(row=row_idx, column=2).value = row['Variant']
    ws.cell(row=row_idx, column=3).value = row['Price']
    ws.cell(row=row_idx, column=4).value = row['Stock']

# Column Width Auto-adjust
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    # Excel limit check (Max width 255 hoti hai)
    ws.column_dimensions[column].width = min(adjusted_width, 100)

wb.save(EXCEL_FILE)
print(f"Excel saved: {EXCEL_FILE}")

# --------------------- EMAIL SENDING ---------------------
print("Preparing Email...")

msg = MIMEMultipart()
msg['From'] = YOUR_EMAIL
msg['To'] = TO_EMAIL
msg['Subject'] = 'Flipkart - Latest Rosier Products List'

body = f"""Hi Automailer,

PFA Flipkart Rosier foods Product Scrapping.
Total Products found: {len(data_rows)}

Thanks!
"""
msg.attach(MIMEText(body, 'plain'))

if os.path.exists(EXCEL_FILE):
    with open(EXCEL_FILE, 'rb') as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename={EXCEL_FILE}')
    msg.attach(part)
else:
    print("Error: Excel file create nahi hui, email attach nahi kar paye.")
    exit()

try:
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(YOUR_EMAIL, APP_PASSWORD)
    server.send_message(msg)
    server.quit()
    print(f"Success! Email sent to {TO_EMAIL}")
except Exception as e:
    print(f"Email Failed: {e}")
